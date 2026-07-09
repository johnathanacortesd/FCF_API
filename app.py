# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any
from pathlib import Path

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias FCF · Realizado por Johnathan Cortés",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-5-nano-2025-08-07"

CONCURRENT_REQUESTS          = 25
SIMILARITY_THRESHOLD_GROUP   = 0.82
SIMILARITY_THRESHOLD_TITULOS = 0.93

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

if 'tokens_input' not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output' not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

# ======================================
# Taxonomía Oficial FCF
# ======================================
TAXONOMIA_FCF = {
    "Institucional": [
        "Comisión Arbitral de la FCF", "82° Congreso Ordinario de la CONMEBOL", "Robo en sede de la FCF en Bogotá",
        "Circulación de procedimiento y requisitos", "Proyecto deportivo del Concejo de Bogotá", "Campeonato Nacional de Primera",
        "Denuncias contra la FCF", "Normativa de la FIFA", "76° Congreso FIFA", "Foto", "Complejo de alto rendimiento de la Selección Colombia",
        "Camiseta Selección Colombia visitante", "Elecciones FCF", "Canción Oficial", "Claro nuevo socio de la Selección Colombia",
        "Proyecto excepción de impuestos", "FCF se une al sello de Carlos Vives", "Investigaciones de la SIC", "Informe de gestión",
        "Dinero por clasificar a fase de grupos", "Patrocinadores FCF", "Reunión de jugadores con directivos", "Camiseta Selección Colombia",
        "Circular informativa FCF", "Assist Card", "Continuidad Ramón Jesurún", "Comisión Arbitral de la Federación Colombiana de Fútbol.",
        "Difútbol", "Comité ejecutivo de la FCF", "Asamblea Ordinaria de la Federación Colombiana de Fútbol", "Alejandro Arteta - Comité ejecutivo FCF",
        "Alianza Fan Zones", "FCF alerta boletería falsa", "Alianza entre Avianca y la FCF", "Comisión del Estatuto del Jugador de la F.C.F",
        "Licencias", "Federación Colombiana de Fútbol", "Bancolombia", "Cámara de Resolución de Disputas de la F.C.F",
        "Importancia de la FCF dentro de la FIFA", "Ingresos FCF", "Tensión en la cúpula del fútbol colombiano", "Ambush marketing (mercadeo parasitario)",
        "FCF lanza al mercado las gafas de la selección", "Logo", "Críticas a la FCF", "Ramón Jesurún", "Imer Machado",
        "Avión oficial de la Selección Colombia", "Ramón Jesurún seguirá siendo el presidente de la FCF", "Lucha contra el contrabando",
        "FCF lanza estrategia para que bares se acrediten como sitios oficiales", "C.D.U de la F.C.F", "Comisión Técnica de la F.C.F",
        "Alianza FCF - Éxito", "Camiseta conmemorativa", "Críticas a Ramón Jesurún", "Comisión Disciplinaria de la FCF", "Comité Ejecutivo de la FCF"
    ],
    "Torneos - Copas - Ligas": [
        "Barranquilla, sede única de la final Sudamericana 2026", "Críticas a árbitros", "Primera C 2026", "Mundial Qatar 2022",
        "Pruebas fisicas para arbitros", "Liga BetPlay Futsal 2026", "Copa América 2028", "Acolfutpro - normativa FIFA",
        "250 equipos llegan al Nacional de Primera C", "Campeonato Sudamericano Sub-17 Fem.", "IFAB", "Zonal nacional",
        "Campeonato Sudamericano Sub-17", "Boletería Mundial 2026", "A 100 días para la Copa del Mundo", "SheBelieves Cup",
        "Técnicos mejor pagos en el Mundial", "Liga Femenina de Fútbol Profesional", "Mobile del FIFAe Nations Cup 2025",
        "La Copa Imposible", "Mundial 2026", "Tiempos de descanso", "Árbitros que están castigados en la Liga BetPlay",
        "Boleteria Liga de Naciones Conmebol en Cali", "Lista de inscritos Liga Betplay 2026.", "Debate sobre la estructura y proyección de la selección Colombia",
        "Federación Colombiana de Fútbol de Salón", "Liga BetPlay 2026", "Boletería Mundial 2026", "Programación Liga de Naciones",
        "Precios boletería Liga de Naciones", "Uso del VAR", "Liga El Dorado", "She Believes Cup 2026", "Mundial de eFootball",
        "Supercopa Juvenil de 2026", "Liga de Naciones Femenina 2025-2026", "Campeonato Nacional de Primera C", "Sorteo del Mundial de la FIFA 2026",
        "Sudamericano sub 20 femenino", "Árbitros colombianos para el Mundial 2026", "Normativa de la Fifa", "Acreditación para medios de comunicación",
        "Designaciones Arbitrales", "Torneo Nacional de Fútbol Sala Femenino 2026", "Zonal Nacional Interligas", "Liga Nacional de Fútbol de Salón 2026"
    ],
    "Selecciones": [
        "Convocatoria S. Mas. de Fútbol Sub 15", "Amistosos antes del Mundial", "Entrenamiento estadio de Techo", "Convocatoria S. Mas. de Fútbol Sub 17",
        "Concentración en Medellín", "Convocatoria S. Fem. de Fútbol Mayores", "S. Mas. de Fútbol Sub 15", "Fredy Hurtado", "Colombia vs Uzbekistan",
        "Néstor Lorenzo", "Calendario de Selecciones Colombia", "Selección Colombia afina detalles rumbo al Mundial 2026", "Convocatoria Mundial 2026",
        "Rueda de prensa Nestor Lorenzo", "Lista de los convocados al Mundial", "Convocatoria S. Mas. de Fútbol Mayores", "Cierre de preparación mundialista",
        "Último amistoso antes de la Copa Mundo", "Plan de la Selección Colombia", "Oliver Glasner", "Convocatoria S. Fem. de Fútbol Sub 17",
        "Sub20 de Fútbol Playa 2026", "Jornada de entrenamiento S. Mayores", "Sede de la Selección Femenina", "Ángelo Marsiglia", "Creación categoría Sub-13 Femenina",
        "Club de la Sele", "Carlos Queiroz", "S. Fem de Fútbol Mayores", "Partido de despedida", "S. Fem de Fútbol Sub 17", "Sedes Selección Colombia",
        "S. Fem de Fútbol Sub 15", "Deivy Mejía, entrenador de la Selección Colombia Sub 15", "Convocatoria doble fecha de amistosos",
        "Copa Mundial Masculina Sub-17", "Selección Femenina Sub 13 de fútbol sala", "Convocatoria S. Fem. de Fútbol Sub 20", "S. Mas. de Fútbol Sub 17",
        "Convocatoria de fútbol playa", "Ciclo de preparación mundialista", "Convocatoria S. Mas. de Fútbol Sub 16", "Convocatoria S. Mas. de Fútbol Sub 19",
        "52 futbolistas en lista para el mundial", "Practica antes del partido de despedida", "Hexagonal final S. Fem de Fútbol Sub 20",
        "S. Mas. de Fútbol Sub 20", "S. Mas. de Fútbol Mayores", "S. Mas. de Fútbol Sub 19", "Copa Mundial Femenina", "S. Fem de Fútbol Sub 20", "Selección colombia"
    ],
    "Gestión": [
        "Avances de obras del Metro", "Avanza recuperación de gramilla del Campín", "Lincencia C", "Acolfutpro - Reformas de estatutos",
        "Pretemporada árbitros del fútbol profesional", "Cancelación registro sindical Acolfutpro", "FCF exalta la gestión de Gianni Infantino al frente de la FIFA",
        "Circular Expedición Pasaportes Deportivos", "Homenaje arbitro Wilmar Roldán", "La FCF pone en marcha las veedurías nacionales", "Programa de detección de talento FCF",
        "FCF envía mensaje de apoyo a México", "Estadios de primera división", "FCF, DIMAYOR y RFEF", "Acreditación y pre-inscripción para la prensa",
        "Regularización entrenadores Futsal y Fútbol Playa", "Seguridad en los estadios", "Curso FIFA RAP Instructores Técnicos de Fútbol 2026",
        "Derechos exclusivos de transmisión de partidos", "Obligatoriedad de licencia C", "Inauguración del Hotel FCF", "Curso RAP", "Los Panitas",
        "Licencia A", "Consolidan al Claro Gaming", "Sesión formativa sobre integridad", "Fútbol con Futuro", "Pruebas de habilitación árbitros categoría A y FIFA",
        "Capacitaciones técnicos y jugadores", "Nuevas reglas para el Mundial 2026", "Licencias para entrenadores", "Licencia B de Entrenador de Arqueros",
        "Fundación HOMI", "Hotel de la Selección Colombia", "Acuerdo laboral en pro de los derechos de los futbolistas", "FCF Camps 2025", "FCF CAMPS 2026",
        "Escarapela FIFA", "Seminario para Entrenadores de Arqueros Juveniles de Élite."
    ],
    "Jugadores": [
        "Aldair Gutiérrez", "Antonio Simancas", "Sebastian Villa", "David Ospina", "Rafael Santos Borré", "Luis Díaz", "Maithe López", "Jefferson Lerma",
        "Samuel Martínez", "Iker Quintero", "Leicy Santos", "Miguel Agámez", "Juan Arizala", "Deiver Machado", "James Rodríguez", "Juan Guillermo Cuadrado",
        "José Escorcia", "Lauren Chamorro López", "Juan Fernando Quintero", "Cristhian Mosquera", "Dávinson Sánchez", "Juan Camilo ´Cucho´ Hernández",
        "Jhon Arias", "Jorge Carrascal", "Falcao García", "Gustavo Puerta", "Johan Mojica", "Sebastián Villa", "Richard Ríos", "Wilmar Barrios",
        "Kevin Castaño", "Daniel Muñoz", "Luis Suárez", "Cristian Borja", "Elkin Arce Mena", "Gorka Abascal", "Jhon Durán", "Jhon Solís", "Jhon Córdoba",
        "Marlyn Sandrid Trujillo Torres", "Dairon Asprilla", "Yoreli Rincón", "Willer Ditta", "Gabriela Rodríguez", "Radamel Falcao", "Hermen Landázuri",
        "Lista de bloqueados", "Juan Camilo Portilla", "Jhon Lucumí", "Santiago Castrillón", "Linda Caicedo"
    ],
    "Entorno": [
        "Alvaro González Alzate", "Amaño de partidos Copa BetPlay", "Amaranto Perea", "Angelica Yulieth Sánchez", "Aniversario del Estadio Metropolitano Roberto Meléndez",
        "Arbitraje Colombiano", "Bobby Moore", "Bogotá aprueba incentivos tributarios", "Bogotá busca brindar incentivos tributarios", "Bogotá FC cumple 23 años",
        "Cartel del fútbol colombiano", "Credito FCF", "Cubrimiento Gol Caracol", "Cubrimiento RCN", "Deportivo Pasto", "Deportivo Pereira",
        "Estadio de Sogamoso", "Exdirectivos", "Faustino \"Tino\" Asprilla", "FCF expresa condolencias a Álvaro González Alzate por el fallecimiento de su hermano",
        "FCF felicita a Atlético Bucaramanga en su aniversario 77", "FCF felicita a Atlético Nacional", "FCF felicita a Bogotá FC", "FCF felicita a Boyacá Chicó",
        "FCF Felicita a FIFA", "FCF felicita a la AUF", "FCF felicita a la Liga de Fútbol de Córdoba", "FCF felicita a la Liga de Fútbol de Quindío",
        "FCF felicita a la Liga de Fútbol de Risaralda", "FCF felicita a la Liga de Fútbol del Valle del Cauca", "FCF felicita a Liga de Fútbol de San Andrés, Providencia, Santa Catalina e Islas",
        "FCF felicita a Liga de Fútbol del Magdalena", "FCF felicita a Llaneros FC", "FCF felicita al Barranquilla FC", "FCF felicita al Unión Magdalena",
        "FCF felicita en el Día Mundial del Entrenador de Fútbol", "FCF lamenta muerte de 'Pocillo' López", "Fotógrafo FCF", "Gafas oficiales FCF",
        "Gobierno Petro", "Hernán Mejía Campuzano", "Homenaje a Santiago Castrillón", "Ímer Machado", "Iván Ramiro Córdoba sueña con la FCF", "Jeison Murillo",
        "Juan Felipe Mejía", "Libro Rastros de una pasión", "Liga Caldense", "Mario Alberto Yepes", "Media FCF", "Miguel Calero", "Noble gesto de Colombia con periodistas",
        "Nuevas prendas antes del mundial", "Nuevas reglas para acabar con la trampa", "Obras en el Estadio Metropolitano", "Rafael Zabaraín",
        "Reemplazar al Deportivo Pereira", "Reinaldo Rueda", "Sección IA", "Sede administrativa de la FCF", "Sede de la Federación Colombiana de Fútbol",
        "Sede partido Inter vs Nacional", "Selección Colombia de Fútbol de Trasplantados", "Shakira y emblemas de la FCF", "Tercera división", "Wilmar Roldán", "Zlatko Dalic"
    ]
}

# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0

if '_emb_cache' not in st.session_state:
    st.session_state['_emb_cache'] = EmbeddingCache()

def get_embedding_cache():
    return st.session_state['_emb_cache']

# ======================================
# CSS Estilizado
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--border:#dadce0;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;
    --accent:#059669;--accent-bg:#ecfdf5;--accent-bdr:#a7f3d0;
    --r2:12px;--shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text',sans-serif;
}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:1.2rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#059669,#047857);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:#047857;font-family:'Roboto Mono',monospace;font-size:0.65rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;text-transform:uppercase;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem;text-align:center;box-shadow:var(--shadow-sm);}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.75rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s2);margin:1rem 0 0.5rem;}
</style>
""", unsafe_allow_html=True)

# ======================================
# Utilidades de Datos
# ======================================
def check_password():
    if st.session_state.get("password_correct", False):
        return True
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            st.markdown("<h3 style='text-align:center;'>Sistema de Análisis FCF</h3>", unsafe_allow_html=True)
            pw = st.text_input("Contraseña", type="password")
            if st.form_submit_button("Ingresar", use_container_width=True):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def load_local_config():
    paths = [Path("Configuracion.xlsx"), Path("configuracion.xlsx"), Path("Config.xlsx"), Path("config.xlsx")]
    for p in paths:
        if p.exists(): return p
    return None

def load_config_maps(config_path):
    try:
        config_sheets = pd.read_excel(config_path, sheet_name=None, engine='openpyxl')
        internet_map = {}
        if 'Internet' in config_sheets:
            internet_map = pd.Series(
                config_sheets['Internet'].iloc[:, 1].values,
                index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
            ).dropna().to_dict()
        return internet_map
    except Exception as e:
        st.warning(f"No se pudo cargar la configuración de mapeos de internet: {e}")
        return {}

def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try: return fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            time.sleep(d); d *= 2

async def acall_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try: return await fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            await asyncio.sleep(d); d *= 2

def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def normalize_title_for_comparison(title):
    if not isinstance(title, str): return ""
    cleaned = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", title).strip()
    if ":" in cleaned:
        parts = cleaned.split(":", 1)
        if len(parts[1].strip()) >= 10: cleaned = parts[1].strip()
    return re.sub(r"\W+", " ", cleaned).lower().strip()

def clean_text(text):
    if not isinstance(text, str): return text
    return text.strip()

def normalizar_tipo_medio_fcf(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    tipo_map = {
        'aire': 'Televisión', 'cable': 'Televisión', 'television': 'Televisión', 'televisión': 'Televisión',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa', 'prensa': 'Prensa',
        'revista': 'Revistas', 'revistas': 'Revistas'
    }
    return tipo_map.get(t, tipo_raw.strip().title())

def buscar_vocero(resumen):
    if not isinstance(resumen, str) or not resumen.strip(): return ""
    res_norm = unidecode(resumen.lower())
    targets = ["ramon jesurun franco", "ramon jesurun", "jesurun"]
    if any(t in res_norm for t in targets):
        return "Ramón Jesurun"
    return ""

def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {r}"[:max_len]

# ======================================
# Estructura DSU para Agrupaciones
# ======================================
class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n): c[self.find(i)].append(i)
        return dict(c)

# ======================================
# Llamadas de Embeddings y Clasificación
# ======================================
def get_embeddings_batch(textos, batch_size=100):
    if not textos: return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing: return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u: st.session_state['tokens_embedding'] += u.get('total_tokens', 0)
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                resultados[oi] = d["embedding"]
                cache.put(textos[oi], d["embedding"])
        except:
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    resultados[oi] = r["data"][0]["embedding"]
                    cache.put(textos[oi], r["data"][0]["embedding"])
                except: pass
    return resultados

async def clasificar_fcf_llm(titulo: str, resumen: str, sem: asyncio.Semaphore) -> dict:
    async with sem:
        # Se estructuran las opciones válidas para inyectar en la instrucción
        taxonomia_instruccion = ""
        for t, subs in TAXONOMIA_FCF.items():
            taxonomia_instruccion += f"- **{t}**: {', '.join(subs[:8])}...\n"

        prompt = (
            f"Eres un analista de inteligencia de medios experto en fútbol y especializado en la Federación Colombiana de Fútbol (FCF).\n"
            f"Tu tarea consiste en analizar la siguiente noticia y clasificar su TEMA, SUBTEMA e IMPACTO reputacional de manera altamente lógica.\n\n"
            f"--- NOTICIA ---\n"
            f"Título: {titulo}\n"
            f"Resumen: {resumen}\n\n"
            f"--- REGLAS DE ETIQUETADO ---\n"
            f"1. **TEMA**: Debe ser EXACTAMENTE una de las siguientes opciones (respeta mayúsculas y acentos):\n"
            f"   - Institucional\n"
            f"   - Torneos - Copas - Ligas\n"
            f"   - Selecciones\n"
            f"   - Gestión\n"
            f"   - Jugadores\n"
            f"   - Entorno\n\n"
            f"2. **SUBTEMA**: Analiza la noticia y asocia un subtema adecuado. Como referencia, aquí tienes algunos ejemplos por tema:\n"
            f"{taxonomia_instruccion}"
            f"   Si la noticia coincide exactamente con uno de estos ejemplos, úsalo. "
            f"Si trata de un asunto específico no listado, genera un subtema nominal muy preciso de 2 a 4 palabras "
            f"(sin verbos conjugados, sin nombres de marcas, respetando acentos y eñes).\n\n"
            f"3. **IMPACTO (Tono)**: Determina el impacto reputacional hacia la FCF:\n"
            f"   - **Positivo**: Si hay exaltación, logros, o reconocimiento directo de la gestión de la FCF.\n"
            f"   - **Negativo**: Críticas directas, denuncias, comentarios negativos o señalamientos hacia la FCF o sus dirigentes.\n"
            f"   - **Neutro**: Información básica sobre partidos de fútbol, listas de convocados, transferencias de jugadores o mención casual sin connotación valorativa.\n\n"
            f"Responde ESTRICTAMENTE en formato JSON con la siguiente estructura:\n"
            f'{{"tema": "TEMA_AQUÍ", "subtema": "SUBTEMA_AQUÍ", "impacto": "Positivo|Negativo|Neutro"}}'
        )

        try:
            resp = await acall_with_retries(
                openai.ChatCompletion.acreate,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=120,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += u.get('prompt_tokens', 0)
                st.session_state['tokens_output'] += u.get('completion_tokens', 0)
            
            res = json.loads(resp.choices[0].message.content)
            tema = str(res.get("tema", "")).strip()
            subtema = str(res.get("subtema", "")).strip()
            impacto = str(res.get("impacto", "")).strip().title()

            if tema not in TAXONOMIA_FCF.keys():
                tema = "Institucional"
            if impacto not in ["Positivo", "Negativo", "Neutro"]:
                impacto = "Neutro"
                
            return {"tema": tema, "subtema": subtema if subtema else "Varios", "impacto": impacto}
        except:
            return {"tema": "Institucional", "subtema": "Varios", "impacto": "Neutro"}

# ======================================
# Proceso de Agrupamiento y Análisis
# ======================================
async def procesar_analisis_fcf(rows, headers, internet_map, pbar):
    n = len(rows)
    pbar.progress(0.05, "Limpiando campos y analizando vocería...")
    
    # 1. Búsquedas Locales y Re-mapeos Obligatorios
    for row in rows:
        # Búsqueda local de vocero en RESUMEN
        resumen_val = row.get("RESUMEN", "") or ""
        row["VOCERO"] = buscar_vocero(str(resumen_val))
        
        # Mapeo y reemplazo de Tipo de Medio
        tm_val = row.get("TIPO DE MEDIO", "") or ""
        tipo_norm = normalizar_tipo_medio_fcf(str(tm_val))
        row["TIPO DE MEDIO"] = tipo_norm
        
        # Buscarv en NOMBRE DE MEDIO si es de Tipo Internet / Online
        if str(tm_val).strip().lower() in ["online", "internet"]:
            nm_val = row.get("NOMBRE DE MEDIO", "") or ""
            nm_clean = str(nm_val).strip().lower()
            if nm_clean in internet_map:
                row["NOMBRE DE MEDIO"] = internet_map[nm_clean]

    # 2. Agrupamiento Lógico de Noticias Similares
    pbar.progress(0.15, "Agrupando por similitud en títulos...")
    dsu = DSU(n)
    titulos = [str(r.get("TÍTULO") or r.get("TITULO", "")).strip() for r in rows]
    norm_titles = [normalize_title_for_comparison(t) for t in titulos]

    for i in range(n):
        if not norm_titles[i]: continue
        for j in range(i + 1, n):
            if not norm_titles[j] or dsu.find(i) == dsu.find(j): continue
            if SequenceMatcher(None, norm_titles[i], norm_titles[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                dsu.union(i, j)

    pbar.progress(0.30, "Generando embeddings para validación semántica...")
    textos = [
        texto_para_embedding(
            str(r.get("TÍTULO") or r.get("TITULO", "")),
            str(r.get("RESUMEN") or "")
        )
        for r in rows
    ]
    embs = get_embeddings_batch(textos)
    validos = [(i, e) for i, e in enumerate(embs) if e is not None]

    if len(validos) >= 2:
        idxs, M = zip(*validos)
        labels = AgglomerativeClustering(
            n_clusters=None,
            distance_threshold=1 - SIMILARITY_THRESHOLD_GROUP,
            metric="cosine",
            linkage="complete"
        ).fit(np.array(M)).labels_
        
        g = defaultdict(list)
        for k, lbl in enumerate(labels):
            g[lbl].append(idxs[k])
            
        for cl in g.values():
            for j in cl[1:]:
                dsu.union(cl[0], j)

    # 3. Procesamiento en Paralelo con Semáforo
    grupos = dsu.grupos(n)
    total_grupos = len(grupos)
    pbar.progress(0.50, f"Clasificando {total_grupos} grupos de noticias con {OPENAI_MODEL_CLASIFICACION}...")

    sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
    tasks = []
    cids = list(grupos.keys())

    for cid in cids:
        rep_idx = grupos[cid][0]
        rep_title = str(rows[rep_idx].get("TÍTULO") or rows[rep_idx].get("TITULO", ""))
        rep_resumen = str(rows[rep_idx].get("RESUMEN") or "")
        tasks.append(clasificar_fcf_llm(rep_title, rep_resumen, sem))

    resultados = []
    for idx, f in enumerate(asyncio.as_completed(tasks)):
        res = await f
        resultados.append(res)
        pbar.progress(0.50 + 0.45 * (idx + 1) / total_grupos, f"Clasificados {idx + 1}/{total_grupos} grupos...")

    rpg = {cids[k]: resultados[k] for k in range(len(cids))}

    # Asignar resultados unificados al grupo completo
    for cid, idxs in grupos.items():
        evaluacion = rpg.get(cid, {"tema": "Institucional", "subtema": "Varios", "impacto": "Neutro"})
        for i in idxs:
            rows[i]["TEMA"] = evaluacion["tema"]
            rows[i]["SUBTEMA"] = evaluacion["subtema"]
            rows[i]["Impacto"] = evaluacion["impacto"]

    pbar.progress(1.0, "Clasificación finalizada")
    return rows

# ======================================
# Generador de Excel de Salida
# ======================================
def generate_fcf_output(rows, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis FCF"

    ws.append(headers)
    font_header = Font(bold=True)
    for idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=idx).font = font_header

    font_hyperlink = Font(color="059669", underline="single")
    align_left = Alignment(horizontal='left')

    for r_idx, row in enumerate(rows, start=2):
        out_row = []
        links = {}
        for c_idx, h in enumerate(headers, start=1):
            val = row.get(h)
            cv = None
            if isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[c_idx] = val["url"]
            else:
                cv = val
            out_row.append(cv)
        ws.append(out_row)

        for c_idx, url in links.items():
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left

        for c_idx, h in enumerate(headers, start=1):
            if str(h).strip().lower() == "fecha":
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                    cell.number_format = 'DD/MM/YYYY'
                elif isinstance(cell.value, pd.Timestamp):
                    cell.value = cell.value.to_pydatetime()
                    cell.number_format = 'DD/MM/YYYY'

    for idx, h in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        h_clean = str(h).strip().lower()
        if "título" in h_clean or "titulo" in h_clean or "resumen" in h_clean:
            ws.column_dimensions[letter].width = 50
        elif "link" in h_clean:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ======================================
# Orquestador del Flujo de Trabajo
# ======================================
async def run_fcf_pipeline(df_file):
    get_embedding_cache().clear()
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    t0 = time.time()

    try:
        openai.api_key = st.secrets["OPENAI_API_KEY"]
        openai.aiosession.set(None)
    except:
        st.error("Error: OPENAI_API_KEY no se encuentra definido.")
        st.stop()

    with st.status("Cargando y procesando archivos de entrada...", expanded=True) as status_step:
        config_path = load_local_config()
        if not config_path:
            st.error("❌ No se encontró el archivo 'Configuracion.xlsx' en la raíz del repositorio.")
            st.stop()
            
        internet_map = load_config_maps(config_path)

        wb_in = load_workbook(df_file, data_only=True)
        sheet = wb_in.active

        # Extraer cabeceras reales
        headers = [str(cell.value) for cell in sheet[1] if cell.value is not None]
        
        # Mapeo de filas del Excel en diccionarios
        rows = []
        for row in sheet.iter_rows(min_row=2):
            if all(c.value is None for c in row): continue
            row_data = {}
            for i, h in enumerate(headers):
                if i < len(row):
                    cell = row[i]
                    val = cell.value
                    url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                    if url:
                        row_data[h] = {"value": val or "Link", "url": url}
                    else:
                        row_data[h] = val
            rows.append(row_data)

        status_step.update(label="✓ Dossier y mapeos cargados correctamente.", state="complete")

    # Validar columnas de entrada críticas
    norm_headers = [h.strip().lower() for h in headers]
    for r_col in ["título", "resumen"]:
        if not any(r_col in nh for nh in norm_headers):
            st.error(f"Falta columna indispensable en el archivo de entrada: {r_col.upper()}")
            st.stop()

    with st.status("Analizando contenido con Inteligencia Artificial...", expanded=True) as status_step:
        pbar = st.progress(0.0)
        rows_processed = await procesar_analisis_fcf(rows, headers, internet_map, pbar)
        status_step.update(label="✓ Clasificación de noticias completada.", state="complete")

    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M

    st.session_state["output_data"] = generate_fcf_output(rows_processed, headers)
    st.session_state["output_filename"] = f"Informe_IA_FCF_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.session_state["processing_complete"] = True
    st.session_state.update({
        "total_rows": len(rows),
        "process_duration": f"{time.time() - t0:.0f}s",
        "process_cost": f"${ci + co + ce:.4f} USD",
        "cache_stats": get_embedding_cache().stats()
    })

# ======================================
# Función Main (Interfaz Streamlit)
# ======================================
def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">⚽</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Noticias FCF · Procesamiento de Reputación</div>
            <div style="font-size:0.75rem; color:#5f6368;">Clasificación de noticias con el modelo gpt-5-nano-2025-08-07</div>
        </div>
        <div class="app-header-badge">FCF IA</div>
    </div>""", unsafe_allow_html=True)

    if not st.session_state.get("processing_complete", False):
        st.markdown('<div class="sec-label">Subir Dossier para Análisis</div>', unsafe_allow_html=True)
        st.markdown("Asegúrate de que tu archivo cuente con las columnas obligatorias: **TÍTULO** y **RESUMEN**.")
        f1 = st.file_uploader("Dossier (.xlsx)", type=["xlsx"], label_visibility="collapsed")

        if st.button("▶ Iniciar Análisis Reputacional", use_container_width=True, type="primary"):
            if not f1:
                st.error("Por favor, sube un archivo antes de continuar.")
            else:
                asyncio.run(run_fcf_pipeline(f1))
                st.rerun()
    else:
        st.markdown(
            '<div style="background-color:var(--accent-bg); border:1px solid var(--accent-bdr); border-radius:12px; padding:1rem; margin-bottom:1rem;">'
            '<h4 style="color:#047857; margin:0;">✓ ¡Procesamiento completado con éxito!</h4>'
            '<p style="margin:0.2rem 0 0 0; font-size:0.85rem;">Las noticias han sido evaluadas y agrupadas lógicamente.</p>'
            '</div>',
            unsafe_allow_html=True
        )

        st.markdown('<div class="sec-label">Métricas de Ejecución</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="metrics-grid">
          <div class="metric-card"><div class="metric-val">{st.session_state.total_rows}</div><div class="metric-lbl">Total Noticias</div></div>
          <div class="metric-card"><div class="metric-val" style="color:#059669;">6</div><div class="metric-lbl">Temas Posibles</div></div>
          <div class="metric-card"><div class="metric-val" style="color:#1a73e8;">{st.session_state.process_duration}</div><div class="metric-lbl">Tiempo de Cómputo</div></div>
          <div class="metric-card"><div class="metric-val" style="color:#d97706;">{st.session_state.process_cost}</div><div class="metric-lbl">Costo Consumo</div></div>
          <div class="metric-card"><div class="metric-val" style="color:#5f6368;">Cache</div><div class="metric-lbl">Estadísticas</div></div>
        </div>""", unsafe_allow_html=True)
        
        st.caption(f"📊 Detalle del cache: {st.session_state.cache_stats}")

        st.markdown('<div class="sec-label">Descargar Resultado</div>', unsafe_allow_html=True)
        st.download_button(
            "⬇ Descargar Informe de Análisis FCF",
            data=st.session_state.output_data,
            file_name=st.session_state.output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        
        if st.button("Procesar un nuevo dossier", use_container_width=True):
            pwd = st.session_state.get("password_correct")
            st.session_state.clear()
            st.session_state.password_correct = pwd
            st.rerun()

    st.markdown(
        '<div style="text-align:center; font-family:\'Roboto Mono\',monospace; font-size:0.65rem; color:#9aa0a6; margin-top:2rem; padding-top:1rem; border-top:1px solid #e8eaed;">'
        'Análisis de Noticias FCF v2.0 · Johnathan Cortés ©'
        '</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
